import numpy as np
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def __grade_absolute(score):
    s = score
    if s > 80:
        return 'A'
    elif s > 70:
        return 'B'
    elif s > 60:
        return 'C'
    elif s >= 50:
        return 'D'
    elif s < 50:
        return 'F'

def __grade_relative(score):
    return 'realative'
        
def process_workbook_absolute_grade(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(5, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        grade = __grade_absolute(score = cell.value)
        grade_cell = sheet.cell(row,4)
        grade_cell.value = grade
    
    values = Reference(sheet,
                min_row=5,
                max_row= sheet.max_row, 
                min_col=3,
                max_col=3)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'j4')

    wb.save(filename)
        
def process_workbook_relative_grade(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    
    for row in range(5, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        grade = __grade_relative(score = cell.value)
        grade_cell = sheet.cell(row,5)
        grade_cell.value = grade
    
    #which values will be used to make graph
    values = Reference(sheet,
                min_row=5,
                max_row= sheet.max_row, 
                min_col=3,
                max_col=3)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'j4')

    wb.save(filename)
        
def write_percentage_grades(filename, rule):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    scores = []
    for row in range(5, sheet.max_row+1):
        name_cell = sheet.cell(row, 2)
        score_cell = sheet.cell(row, 3)
        name = name_cell.value
        score = score_cell.value
        scores.append([name, score, 'I'])

    data = np.array(scores)
    data = data[data[:,1].argsort()]  # sort on column 1
    data = data[::-1]  # sort on column 1
    
    
    rule = np.array(rule)
    total = rule[:,0].astype(np.float)
    
    
    if total.sum() != 1.0:
        print("Please give valid rule in which percentages sum = 100%")
        return

    base = 0
    for r in rule:
        percentage = float(r[0])
        grade = r[1]
        limit = base + percentage
        baseN = round(data.shape[0] * base)
        limitN = round(data.shape[0] * limit)
        

        data[baseN:limitN,2]=grade
        base = limit


    print('data grades assigned:')
    print(data)

def print_relative_grades(filename, rule):

    #main work:
    #implement relative, k-means grading and compare them and share insights
    #tips = gaps finding and gaps dealing.

    #simple implement 30 mins
    #plan:
    #phase 1:
    #get marks list, sort them find mean, std. out put

    #phase 2:
    #divide the list in provided number of division with max = std+2m, min =std-2m, divisons = no of grades
    #out put divisons
    #assign provided grades

    #phase 3:
    #write the grades to excel file

    #phase 4:
    #write the grades to excel file in order they were written before

    #phase 5:
    #get rules from a excel sheet of grades. an email for students for suggestions of
    #   mathematical models in grading. i.e normal dist, heuristic, curve, cone, sin, derivative, integration, random etc.
    
    #phase 6:
    #write the grades to excel file in order they were written before


    #later add an id from 1 to n before sorting so that we can arrange them on same 

    #later provide time complexity big O and normal running times on 100 , 1000, 100 0000000 students data etc. i.e 1 s, 4, sec, 100 sec etc.

    #siple thoughts 1 hr
    #sd #mean
    #assign 1 grade, 2 grades, 3 grades etc...
    #mean-3sd mean-2sd mean-1sd mean  mean+1sd mean+2sd mean+3sd

    #<1 please provide some number
    #1 grade = all curve 1 grade
    #2 grade = above mean, below mean
    #3 grade = 
    
    #total curve = scores = mean-2d to mean+2d
    # i.e 10 to 90

    # grades = 2
    # so base = 10 limit = 90
    # 90-10 = 80 , 80/2grades = 40, so 0 to 10base+40 then 50 to end
    
    # grades = 3 F,B,A
    # so base = 20 limit = 80
    # 80-20 = 80 , 60/3 grades = 20, so 0to20to40 = F, 40to60=B, 60to80to100=A

    #and so on

    #4 grades
    #    F        D      C       B        A
    # mean-2sd mean-1sd mean  mean+1sd mean+2sd

    #5 grades
    #    F        D      C       B        A
    # mean-2sd mean-1sd mean  mean+1sd mean+2sd

    return ''


rule = [[0.05,'A+'], [0.05,'A'], [0.05,'A-'], 
        [0.10,'B+'], [0.10,'B'], [0.10,'B-'], 
        [0.10,'C+'], [0.10,'C'], [0.10,'C-'], 
        [0.10,'D+'], [0.05,'D'], [0.05,'D-'],
        [0.05,'F']]

print_relative_grades('marks_to_grades.xlsx', rule)
print('--------done')










   #data[:,2]='NG' #data[all rows, 2nd column of all]='NG'
    #data[:5,2]='NG' #data[0 to 5 rows, 2nd column of all]='NG'
    #data[5:,2]='YG' #data[5 to end rows, 2nd column of all]='NG'
    
    #print(round(total.sum()))
    #print('data old:')
    #print(data)

    
        # a = data.shape[0] * base
        # b = data.shape[0] * limit
        #base , round, int compare values
        #print(str(round(a,2)) + ',' + str(round(b,2)) + ' ' + str(int(a))+ ',' +str(int(b)) + ' ' + str(round(b)) + ',' + str(round(b)))

        #print('base:'+str(round(baseN,2)) + ' limit:' + str(round(limitN,2)))
        #print('r bas:'+str(round(baseN)) + ' r limt:' + str(round(limitN)))

    #print(rule[:,0].shape)
    #print(np.array([1,2,3]).sum())
    #print(np.array(rule[:,0]).sum())
    #print(data)

    #data[][2] = 'A'
    #data is sorted
    #base = 0
    #for r in rule
    #   data[base to r[i]] = r[1] #assigning grades from 0 to 10% then 10 to 20% and so on till 100%
    #   base = r[i]
    # I did some coding in python. I was trying to implement relative grading. Python skills were weak. So I will revise python in Sololearn app and code tomorrow IA.
  